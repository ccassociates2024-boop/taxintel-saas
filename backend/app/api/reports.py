"""PDF and Excel report generation for client tax summaries."""
from __future__ import annotations

import datetime
from io import BytesIO
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)
from sqlalchemy.orm import Session

from app.api.deps import get_current_tenant
from app.db.models import Tenant
from app.db.session import get_db
from app.repositories import ClientRepository, TaxRepository

router = APIRouter(prefix="/reports", tags=["reports"])

# ── Brand colours ──────────────────────────────────────────────────────────
INDIGO = colors.HexColor("#4F46E5")
SLATE  = colors.HexColor("#1E293B")
MUTED  = colors.HexColor("#64748B")
GREEN  = colors.HexColor("#16A34A")
RED    = colors.HexColor("#DC2626")
BG     = colors.HexColor("#F1F5F9")


def _money(val) -> str:
    try:
        return f"₹{float(val):,.2f}"
    except Exception:
        return str(val)


@router.get("/{client_id}/pdf")
def pdf_report(
    client_id: UUID,
    assessment_year: str = "2024-25",
    tenant: Tenant = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    client = ClientRepository(db).get_for_tenant(tenant.id, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    repo = TaxRepository(db)
    ais      = repo.latest_ais(tenant.id, client_id)
    form26   = repo.latest_26as(tenant.id, client_id)
    comp     = repo.latest_computation(tenant.id, client_id)
    recs     = repo.latest_recommendations(tenant.id, client_id)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=20*mm, bottomMargin=20*mm,
    )

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], textColor=INDIGO, fontSize=18, spaceAfter=4)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], textColor=SLATE, fontSize=12, spaceAfter=2)
    body = ParagraphStyle("Body", parent=styles["Normal"], textColor=MUTED, fontSize=9, spaceAfter=2)
    badge_h = ParagraphStyle("Badge", fontSize=8, textColor=colors.white, alignment=1)

    story = []

    # ── Header ─────────────────────────────────────────────────────────────
    story.append(Paragraph("TaxIntel AI — Tax Intelligence Report", h1))
    story.append(Paragraph(
        f"Firm: <b>{tenant.trade_name}</b> &nbsp;|&nbsp; "
        f"AY: <b>{assessment_year}</b> &nbsp;|&nbsp; "
        f"Generated: {datetime.date.today().strftime('%d %b %Y')}",
        body,
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=INDIGO, spaceAfter=6))

    # ── Client Info ────────────────────────────────────────────────────────
    story.append(Paragraph("Client Information", h2))
    ci_data = [
        ["Name", client.full_name, "PAN", client.pan],
        ["Type", client.client_type, "Status", client.residential_status],
        ["Email", client.email or "—", "Phone", client.phone or "—"],
    ]
    ci_table = Table(ci_data, colWidths=[35*mm, 65*mm, 25*mm, 45*mm])
    ci_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), BG),
        ("BACKGROUND", (2, 0), (2, -1), BG),
        ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",   (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("TEXTCOLOR",  (0, 0), (-1, -1), SLATE),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, BG]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
    ]))
    story.append(ci_table)
    story.append(Spacer(1, 8))

    # ── Tax Computation ────────────────────────────────────────────────────
    if comp:
        story.append(Paragraph("Tax Computation Summary", h2))
        regime_color = GREEN if comp.recommended_regime == "NEW" else INDIGO
        tc_data = [
            ["Old Regime Tax",  _money(comp.old_regime_tax), "New Regime Tax", _money(comp.new_regime_tax)],
            ["Recommended",     comp.recommended_regime,    "Health Score",   f"{comp.health_score} / 100"],
            ["Tax Payable",     _money(comp.tax_payable),   "Refund Est.",    _money(comp.refund_estimate)],
        ]
        tc_table = Table(tc_data, colWidths=[40*mm, 55*mm, 35*mm, 40*mm])
        tc_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), BG),
            ("BACKGROUND", (2, 0), (2, -1), BG),
            ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME",   (2, 0), (2, -1), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("TEXTCOLOR",  (0, 0), (-1, -1), SLATE),
            ("TEXTCOLOR",  (1, 1), (1, 1), regime_color),
            ("FONTNAME",   (1, 1), (1, 1), "Helvetica-Bold"),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, BG]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ]))
        story.append(tc_table)
        story.append(Spacer(1, 8))

    # ── AIS Summary ────────────────────────────────────────────────────────
    if ais:
        story.append(Paragraph("Annual Information Statement (AIS)", h2))
        ais_raw = ais.raw_json or {}
        ais_rows = [
            ["Salary Income",    _money(ais_raw.get("salary", 0))],
            ["Interest Income",  _money(ais_raw.get("interest_income", 0))],
            ["Dividend Income",  _money(ais_raw.get("dividend_income", 0))],
            ["Capital Gains",    _money(ais_raw.get("capital_gains", 0))],
            ["TDS / TCS",        _money(ais_raw.get("tds_tcs", 0))],
        ]
        ais_table = Table([["Category", "Amount"]] + ais_rows,
                          colWidths=[80*mm, 90*mm])
        ais_table.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), INDIGO),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 9),
            ("TEXTCOLOR",     (0, 1), (-1, -1), SLATE),
            ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, BG]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ]))
        story.append(ais_table)
        story.append(Spacer(1, 8))

    # ── Recommendations ────────────────────────────────────────────────────
    if recs:
        story.append(Paragraph("AI Tax Recommendations", h2))
        rec_data = [["Priority", "Category", "Title", "Est. Savings"]]
        priority_colors = {"HIGH": RED, "MEDIUM": colors.HexColor("#D97706"), "LOW": GREEN}
        for rec in recs[:6]:
            rec_data.append([
                rec.priority,
                rec.category,
                rec.title,
                _money(rec.estimated_savings),
            ])
        rec_table = Table(rec_data, colWidths=[22*mm, 28*mm, 90*mm, 30*mm])
        rec_table.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), SLATE),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("TEXTCOLOR",     (0, 1), (-1, -1), SLATE),
            ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, BG]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("WORDWRAP",      (2, 1), (2, -1), True),
        ]))
        story.append(rec_table)
        story.append(Spacer(1, 8))

    # ── Footer ─────────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=MUTED))
    story.append(Paragraph(
        f"Confidential — {tenant.trade_name} | GSTIN: {tenant.gstin or 'N/A'} | "
        "Generated by TaxIntel AI. Not a substitute for professional CA advice.",
        ParagraphStyle("Footer", fontSize=7, textColor=MUTED),
    ))

    doc.build(story)
    buffer.seek(0)
    fname = f"taxintel-{client.full_name.replace(' ', '-')}-{assessment_year}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{client_id}/excel")
def excel_report(
    client_id: UUID,
    assessment_year: str = "2024-25",
    tenant: Tenant = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    client = ClientRepository(db).get_for_tenant(tenant.id, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    repo = TaxRepository(db)
    ais    = repo.latest_ais(tenant.id, client_id)
    form26 = repo.latest_26as(tenant.id, client_id)
    comp   = repo.latest_computation(tenant.id, client_id)
    recs   = repo.latest_recommendations(tenant.id, client_id)

    wb = Workbook()

    # ── Colour palette ────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="4F46E5")
    SUB_FILL  = PatternFill("solid", fgColor="E0E7FF")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    BOLD      = Font(bold=True, size=10)
    NORMAL    = Font(size=10)

    def _hdr(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center")
        return c

    def _sub(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = SUB_FILL; c.font = BOLD
        return c

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32

    _hdr(ws, 1, 1, "TaxIntel AI — Tax Intelligence Report")
    ws.merge_cells("A1:B1")
    ws.cell(row=2, column=1, value=f"Firm: {tenant.trade_name}").font = BOLD
    ws.cell(row=2, column=2, value=f"AY: {assessment_year}").font = BOLD
    ws.cell(row=3, column=1, value=f"Generated: {datetime.date.today()}").font = NORMAL

    _sub(ws, 5, 1, "Client Information")
    rows = [
        ("Full Name", client.full_name),
        ("PAN", client.pan),
        ("Email", client.email or "—"),
        ("Phone", client.phone or "—"),
        ("Type", client.client_type),
        ("Residential Status", client.residential_status),
    ]
    for i, (k, v) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=k).font = BOLD
        ws.cell(row=i, column=2, value=v).font = NORMAL

    if comp:
        r = 13
        _sub(ws, r, 1, "Tax Computation")
        comp_rows = [
            ("Old Regime Tax", float(comp.old_regime_tax)),
            ("New Regime Tax", float(comp.new_regime_tax)),
            ("Recommended Regime", comp.recommended_regime),
            ("Tax Payable", float(comp.tax_payable)),
            ("Refund Estimate", float(comp.refund_estimate)),
            ("Health Score", f"{comp.health_score}/100"),
        ]
        for i, (k, v) in enumerate(comp_rows, start=r+1):
            ws.cell(row=i, column=1, value=k).font = BOLD
            ws.cell(row=i, column=2, value=v).font = NORMAL

    # ── Sheet 2: AIS ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("AIS Data")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20
    _hdr(ws2, 1, 1, "Annual Information Statement (AIS)")
    ws2.merge_cells("A1:B1")
    ws2.cell(row=2, column=1, value="AY").font = BOLD
    ws2.cell(row=2, column=2, value=assessment_year)
    _hdr(ws2, 3, 1, "Category")
    _hdr(ws2, 3, 2, "Amount (INR)")
    if ais:
        raw = ais.raw_json or {}
        fields = [
            ("Salary", raw.get("salary", 0)),
            ("Interest Income", raw.get("interest_income", 0)),
            ("Dividend Income", raw.get("dividend_income", 0)),
            ("Capital Gains", raw.get("capital_gains", 0)),
            ("TDS / TCS", raw.get("tds_tcs", 0)),
            ("Foreign Remittance", raw.get("foreign_remittance", 0)),
            ("High Value Transactions", raw.get("high_value_transactions", 0)),
        ]
        for i, (k, v) in enumerate(fields, start=4):
            ws2.cell(row=i, column=1, value=k).font = BOLD
            ws2.cell(row=i, column=2, value=float(v)).number_format = "#,##0.00"
    else:
        ws2.cell(row=4, column=1, value="No AIS uploaded yet").font = NORMAL

    # ── Sheet 3: Recommendations ──────────────────────────────────────────
    ws3 = wb.create_sheet("Recommendations")
    for col, width in [("A", 12), ("B", 16), ("C", 45), ("D", 20)]:
        ws3.column_dimensions[col].width = width
    _hdr(ws3, 1, 1, "Priority")
    _hdr(ws3, 1, 2, "Category")
    _hdr(ws3, 1, 3, "Title")
    _hdr(ws3, 1, 4, "Est. Savings (INR)")
    if recs:
        for i, rec in enumerate(recs, start=2):
            ws3.cell(row=i, column=1, value=rec.priority).font = BOLD
            ws3.cell(row=i, column=2, value=rec.category)
            ws3.cell(row=i, column=3, value=rec.title)
            ws3.cell(row=i, column=4, value=float(rec.estimated_savings)).number_format = "#,##0.00"
    else:
        ws3.cell(row=2, column=1, value="No recommendations yet").font = NORMAL

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    fname = f"taxintel-{client.full_name.replace(' ', '-')}-{assessment_year}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
